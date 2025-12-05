# app.py - Flask Backend with MongoDB
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from flask_pymongo import PyMongo
from werkzeug.security import generate_password_hash, check_password_hash
from bson.objectid import ObjectId
from datetime import datetime
import jwt
from functools import wraps
import os
from dotenv import load_dotenv
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
import requests
import json
from datetime import timedelta

# Load environment variables
load_dotenv()

app = Flask(__name__)
CORS(app, resources={r"/api/*": {"origins": "*"}}, supports_credentials=True)

# Configuration
app.config['MONGO_URI'] = os.getenv('MONGO_URI', 'mongodb://localhost:27017/streamline_db')
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'your-secret-key-change-this')

mongo = PyMongo(app)

# Token verification decorator
def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization')
        
        if not token:
            return jsonify({'message': 'Token is missing!'}), 401
        
        try:
            token = token.split(' ')[1]  # Remove 'Bearer ' prefix
            data = jwt.decode(token, app.config['SECRET_KEY'], algorithms=['HS256'])
            current_user = mongo.db.users.find_one({'_id': ObjectId(data['user_id'])})
        except:
            return jsonify({'message': 'Token is invalid!'}), 401
        
        return f(current_user, *args, **kwargs)
    
    return decorated

# Auth Routes
@app.route('/api/auth/register', methods=['POST'])
def register():
    data = request.get_json()
    
    email = data.get('email')
    password = data.get('password')
    name = data.get('name', '')
    mobile = data.get('mobile', '')
    
    if not email or not password:
        return jsonify({'message': 'Email and password required'}), 400
    
    if not name:
        return jsonify({'message': 'Name is required'}), 400
    
    # Check if user exists
    if mongo.db.users.find_one({'email': email}):
        return jsonify({'message': 'User already exists'}), 400
    
    # Create user with extended profile
    hashed_password = generate_password_hash(password)
    user_id = mongo.db.users.insert_one({
        'email': email,
        'password': hashed_password,
        'name': name,
        'mobile': mobile,
        'avatar': '',
        'timezone': 'UTC',
        'theme': 'light',
        'notifications': {
            'email': True,
            'push': True,
            'reminders': True
        },
        'preferences': {
            'default_priority': 'medium',
            'default_category': 'personal',
            'auto_archive': False,
            'show_completed': True
        },
        'api_settings': {
            'ai_provider': 'gemini',  # gemini, openai, claude, custom
            'gemini_api_key': '',
            'openai_api_key': '',
            'claude_api_key': '',
            'custom_api_endpoint': '',
            'custom_api_key': ''
        },
        'created_at': datetime.utcnow(),
        'updated_at': datetime.utcnow()
    }).inserted_id
    
    # Generate token
    token = jwt.encode({
        'user_id': str(user_id)
    }, app.config['SECRET_KEY'], algorithm='HS256')
    
    return jsonify({
        'message': 'User created successfully',
        'token': token,
        'user': {
            'id': str(user_id),
            'email': email,
            'name': name,
            'mobile': mobile
        }
    }), 201

@app.route('/api/auth/login', methods=['POST'])
def login():
    data = request.get_json()
    
    email = data.get('email')
    password = data.get('password')
    
    if not email or not password:
        return jsonify({'message': 'Email and password required'}), 400
    
    user = mongo.db.users.find_one({'email': email})
    
    if not user or not check_password_hash(user['password'], password):
        return jsonify({'message': 'Invalid credentials'}), 401
    
    token = jwt.encode({
        'user_id': str(user['_id'])
    }, app.config['SECRET_KEY'], algorithm='HS256')
    
    return jsonify({
        'message': 'Login successful',
        'token': token,
        'user': {
            'id': str(user['_id']),
            'email': user['email'],
            'name': user.get('name', ''),
            'mobile': user.get('mobile', ''),
            'avatar': user.get('avatar', ''),
            'theme': user.get('theme', 'light'),
            'preferences': user.get('preferences', {}),
            'notifications': user.get('notifications', {})
        }
    }), 200

# User Profile Routes
@app.route('/api/user/profile', methods=['GET'])
@token_required
def get_user_profile(current_user):
    user_data = {
        'id': str(current_user['_id']),
        'email': current_user['email'],
        'name': current_user.get('name', ''),
        'mobile': current_user.get('mobile', ''),
        'avatar': current_user.get('avatar', ''),
        'timezone': current_user.get('timezone', 'UTC'),
        'theme': current_user.get('theme', 'light'),
        'notifications': current_user.get('notifications', {
            'email': True,
            'push': True,
            'reminders': True
        }),
        'preferences': current_user.get('preferences', {
            'default_priority': 'medium',
            'default_category': 'personal',
            'auto_archive': False,
            'show_completed': True
        }),
        'api_settings': current_user.get('api_settings', {
            'ai_provider': 'gemini',
            'gemini_api_key': '',
            'openai_api_key': '',
            'claude_api_key': '',
            'custom_api_endpoint': '',
            'custom_api_key': ''
        }),
        'created_at': current_user['created_at'].isoformat() if current_user.get('created_at') else '',
        'updated_at': current_user.get('updated_at', datetime.utcnow()).isoformat()
    }
    return jsonify(user_data), 200

@app.route('/api/user/profile', methods=['PUT'])
@token_required
def update_user_profile(current_user):
    data = request.get_json()
    
    update_data = {'updated_at': datetime.utcnow()}
    
    # Basic profile fields
    if 'name' in data:
        update_data['name'] = data['name']
    if 'mobile' in data:
        update_data['mobile'] = data['mobile']
    if 'avatar' in data:
        update_data['avatar'] = data['avatar']
    if 'timezone' in data:
        update_data['timezone'] = data['timezone']
    if 'theme' in data:
        update_data['theme'] = data['theme']
    
    # Nested objects
    if 'notifications' in data:
        update_data['notifications'] = data['notifications']
    if 'preferences' in data:
        update_data['preferences'] = data['preferences']
    if 'api_settings' in data:
        update_data['api_settings'] = data['api_settings']
    
    # Update user
    mongo.db.users.update_one(
        {'_id': current_user['_id']},
        {'$set': update_data}
    )
    
    # Log activity
    mongo.db.activities.insert_one({
        'user_id': str(current_user['_id']),
        'type': 'profile_updated',
        'description': 'Updated profile settings',
        'created_at': datetime.utcnow()
    })
    
    return jsonify({'message': 'Profile updated successfully'}), 200

@app.route('/api/user/change-password', methods=['POST'])
@token_required
def change_password(current_user):
    data = request.get_json()
    
    current_password = data.get('current_password')
    new_password = data.get('new_password')
    
    if not current_password or not new_password:
        return jsonify({'message': 'Current and new password required'}), 400
    
    # Verify current password
    if not check_password_hash(current_user['password'], current_password):
        return jsonify({'message': 'Current password is incorrect'}), 400
    
    # Update password
    hashed_new_password = generate_password_hash(new_password)
    mongo.db.users.update_one(
        {'_id': current_user['_id']},
        {'$set': {
            'password': hashed_new_password,
            'updated_at': datetime.utcnow()
        }}
    )
    
    # Log activity
    mongo.db.activities.insert_one({
        'user_id': str(current_user['_id']),
        'type': 'password_changed',
        'description': 'Changed account password',
        'created_at': datetime.utcnow()
    })
    
    return jsonify({'message': 'Password changed successfully'}), 200

# Todo Routes
@app.route('/api/todos', methods=['GET'])
@token_required
def get_todos(current_user):
    todos = list(mongo.db.todos.find({'user_id': str(current_user['_id'])}))
    
    # Convert ObjectId to string
    for todo in todos:
        todo['_id'] = str(todo['_id'])
        todo['created_at'] = todo['created_at'].isoformat()
    
    return jsonify(todos), 200

@app.route('/api/todos', methods=['POST'])
@token_required
def create_todo(current_user):
    data = request.get_json()
    
    todo = {
        'text': data.get('text'),
        'completed': False,
        'priority': data.get('priority', 'medium'),
        'category': data.get('category', 'personal'),
        'user_id': str(current_user['_id']),
        'created_at': datetime.utcnow(),
        'time_spent': 0,  # in minutes
        'started_at': None,
        'completed_at': None,
        'estimated_time': data.get('estimated_time', 0)  # in minutes
    }
    
    result = mongo.db.todos.insert_one(todo)
    todo['_id'] = str(result.inserted_id)
    todo['created_at'] = todo['created_at'].isoformat()
    
    return jsonify(todo), 201

@app.route('/api/todos/<todo_id>', methods=['PUT'])
@token_required
def update_todo(current_user, todo_id):
    data = request.get_json()
    
    # Verify todo belongs to user
    todo = mongo.db.todos.find_one({
        '_id': ObjectId(todo_id),
        'user_id': str(current_user['_id'])
    })
    
    if not todo:
        return jsonify({'message': 'Todo not found'}), 404
    
    update_data = {}
    if 'text' in data:
        update_data['text'] = data['text']
    if 'completed' in data:
        update_data['completed'] = data['completed']
    if 'priority' in data:
        update_data['priority'] = data['priority']
    if 'category' in data:
        update_data['category'] = data['category']
    
    mongo.db.todos.update_one(
        {'_id': ObjectId(todo_id)},
        {'$set': update_data}
    )
    
    # Log activity for completed tasks
    if 'completed' in data and data['completed'] and not todo.get('completed'):
        # Task completed
        mongo.db.activities.insert_one({
            'user_id': str(current_user['_id']),
            'type': 'task_completed',
            'description': f'Completed task: {todo["text"]}',
            'task_id': todo_id,
            'created_at': datetime.utcnow()
        })
        update_data['completed_at'] = datetime.utcnow()
    elif 'completed' in data and not data['completed'] and todo.get('completed'):
        # Task reopened
        mongo.db.activities.insert_one({
            'user_id': str(current_user['_id']),
            'type': 'task_reopened',
            'description': f'Reopened task: {todo["text"]}',
            'task_id': todo_id,
            'created_at': datetime.utcnow()
        })
        update_data['completed_at'] = None

    return jsonify({'message': 'Todo updated successfully'}), 200

@app.route('/api/todos/<todo_id>', methods=['DELETE'])
@token_required
def delete_todo(current_user, todo_id):
    result = mongo.db.todos.delete_one({
        '_id': ObjectId(todo_id),
        'user_id': str(current_user['_id'])
    })
    
    if result.deleted_count == 0:
        return jsonify({'message': 'Todo not found'}), 404
    
    # Delete associated comments
    mongo.db.comments.delete_many({'todo_id': todo_id})
    
    return jsonify({'message': 'Todo deleted successfully'}), 200

# Get single todo with details
@app.route('/api/todos/<todo_id>', methods=['GET'])
@token_required
def get_todo(current_user, todo_id):
    todo = mongo.db.todos.find_one({
        '_id': ObjectId(todo_id),
        'user_id': str(current_user['_id'])
    })
    
    if not todo:
        return jsonify({'message': 'Todo not found'}), 404
    
    todo['_id'] = str(todo['_id'])
    todo['created_at'] = todo['created_at'].isoformat()
    if todo.get('completed_at'):
        todo['completed_at'] = todo['completed_at'].isoformat()
    if todo.get('started_at'):
        todo['started_at'] = todo['started_at'].isoformat()
    
    return jsonify(todo), 200

# Start/Stop time tracking
@app.route('/api/todos/<todo_id>/time', methods=['POST'])
@token_required
def toggle_time_tracking(current_user, todo_id):
    data = request.get_json()
    action = data.get('action')  # 'start' or 'stop'
    
    todo = mongo.db.todos.find_one({
        '_id': ObjectId(todo_id),
        'user_id': str(current_user['_id'])
    })
    
    if not todo:
        return jsonify({'message': 'Todo not found'}), 404
    
    if action == 'start':
        if todo.get('started_at'):
            return jsonify({'message': 'Timer already running'}), 400
        
        mongo.db.todos.update_one(
            {'_id': ObjectId(todo_id)},
            {'$set': {'started_at': datetime.utcnow()}}
        )
        
        # Log activity
        mongo.db.activities.insert_one({
            'user_id': str(current_user['_id']),
            'type': 'timer_started',
            'description': f'Started working on: {todo["text"]}',
            'task_id': todo_id,
            'created_at': datetime.utcnow()
        })
        
        return jsonify({'message': 'Timer started'}), 200
    
    elif action == 'stop':
        if not todo.get('started_at'):
            return jsonify({'message': 'Timer not running'}), 400
        
        started_at = todo['started_at']
        time_diff = datetime.utcnow() - started_at
        minutes_spent = int(time_diff.total_seconds() / 60)
        
        current_time_spent = todo.get('time_spent', 0)
        new_time_spent = current_time_spent + minutes_spent
        
        mongo.db.todos.update_one(
            {'_id': ObjectId(todo_id)},
            {
                '$set': {'time_spent': new_time_spent},
                '$unset': {'started_at': ''}
            }
        )
        
        # Log activity
        mongo.db.activities.insert_one({
            'user_id': str(current_user['_id']),
            'type': 'timer_stopped',
            'description': f'Worked {minutes_spent} minutes on: {todo["text"]}',
            'task_id': todo_id,
            'time_spent': minutes_spent,
            'created_at': datetime.utcnow()
        })
        
        return jsonify({
            'message': 'Timer stopped',
            'time_spent': minutes_spent,
            'total_time': new_time_spent
        }), 200
    
    return jsonify({'message': 'Invalid action'}), 400

# Comment Routes
@app.route('/api/todos/<todo_id>/comments', methods=['GET'])
@token_required
def get_comments(current_user, todo_id):
    # Verify todo belongs to user
    todo = mongo.db.todos.find_one({
        '_id': ObjectId(todo_id),
        'user_id': str(current_user['_id'])
    })
    
    if not todo:
        return jsonify({'message': 'Todo not found'}), 404
    
    comments = list(mongo.db.comments.find({'todo_id': todo_id}).sort('created_at', -1))
    
    for comment in comments:
        comment['_id'] = str(comment['_id'])
        comment['created_at'] = comment['created_at'].isoformat()
    
    return jsonify(comments), 200

@app.route('/api/todos/<todo_id>/comments', methods=['POST'])
@token_required
def add_comment(current_user, todo_id):
    # Verify todo belongs to user
    todo = mongo.db.todos.find_one({
        '_id': ObjectId(todo_id),
        'user_id': str(current_user['_id'])
    })
    
    if not todo:
        return jsonify({'message': 'Todo not found'}), 404
    
    data = request.get_json()
    
    comment = {
        'todo_id': todo_id,
        'text': data.get('text'),
        'user_email': current_user['email'],
        'created_at': datetime.utcnow()
    }
    
    result = mongo.db.comments.insert_one(comment)
    comment['_id'] = str(result.inserted_id)
    comment['created_at'] = comment['created_at'].isoformat()
    
    return jsonify(comment), 201

@app.route('/api/todos/<todo_id>/comments/<comment_id>', methods=['DELETE'])
@token_required
def delete_comment(current_user, todo_id, comment_id):
    result = mongo.db.comments.delete_one({
        '_id': ObjectId(comment_id),
        'todo_id': todo_id
    })
    
    if result.deleted_count == 0:
        return jsonify({'message': 'Comment not found'}), 404
    
    return jsonify({'message': 'Comment deleted successfully'}), 200

# Activity Feed
@app.route('/api/activities', methods=['GET'])
@token_required
def get_activities(current_user):
    limit = int(request.args.get('limit', 50))
    activities = list(mongo.db.activities.find(
        {'user_id': str(current_user['_id'])}
    ).sort('created_at', -1).limit(limit))
    
    for activity in activities:
        activity['_id'] = str(activity['_id'])
        activity['created_at'] = activity['created_at'].isoformat()
    
    return jsonify(activities), 200

@app.route('/api/health', methods=['GET'])
def health_check():
    return jsonify({'status': 'healthy'}), 200

# Debug endpoint to check user settings
@app.route('/api/debug/user-settings', methods=['GET'])
@token_required
def debug_user_settings(current_user):
    api_settings = current_user.get('api_settings', {})
    ai_provider = api_settings.get('ai_provider', 'gemini')
    
    # Get the actual API key that would be used
    if ai_provider == 'gemini':
        api_key = api_settings.get('gemini_api_key', '') or os.getenv('GEMINI_API_KEY')
    elif ai_provider == 'openai':
        api_key = api_settings.get('openai_api_key', '') or os.getenv('OPENAI_API_KEY')
    elif ai_provider == 'claude':
        api_key = api_settings.get('claude_api_key', '') or os.getenv('CLAUDE_API_KEY')
    elif ai_provider == 'custom':
        api_key = api_settings.get('custom_api_key', '')
    else:
        api_key = None
    
    return jsonify({
        'user_id': str(current_user['_id']),
        'email': current_user['email'],
        'api_settings': api_settings,
        'ai_provider': ai_provider,
        'has_api_key': bool(api_key),
        'api_key_length': len(api_key) if api_key else 0,
        'api_key_preview': api_key[:10] + '...' if api_key and len(api_key) > 10 else api_key,
        'has_env_gemini_key': bool(os.getenv('GEMINI_API_KEY')),
        'env_gemini_key_length': len(os.getenv('GEMINI_API_KEY', '')),
        'gemini_url': f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={'[KEY]' if api_key else '[NO_KEY]'}"
    }), 200

# Test AI API endpoint
@app.route('/api/debug/test-ai', methods=['POST'])
@token_required
def test_ai_api(current_user):
    api_settings = current_user.get('api_settings', {})
    ai_provider = api_settings.get('ai_provider', 'gemini')
    
    # Get API key
    if ai_provider == 'gemini':
        api_key = api_settings.get('gemini_api_key', '') or os.getenv('GEMINI_API_KEY')
    else:
        return jsonify({'message': 'Only Gemini testing supported for now'}), 400
    
    if not api_key:
        return jsonify({'message': 'No API key found'}), 400
    
    # Simple test prompt
    test_prompt = "Say 'Hello, this is a test response' in JSON format: {\"message\": \"your response\"}"
    
    try:
        url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
        payload = {
            "contents": [{
                "parts": [{
                    "text": test_prompt
                }]
            }]
        }
        
        response = requests.post(url, json=payload)
        
        return jsonify({
            'status_code': response.status_code,
            'response_text': response.text[:1000],  # First 1000 chars
            'url': url.replace(api_key, '[HIDDEN_KEY]'),
            'success': response.status_code == 200
        }), 200
        
    except Exception as e:
        return jsonify({
            'error': str(e),
            'success': False
        }), 200

# Analytics and Stats
@app.route('/api/analytics/stats', methods=['GET'])
@token_required
def get_user_stats(current_user):
    user_id = str(current_user['_id'])
    
    # Basic counts
    total_todos = mongo.db.todos.count_documents({'user_id': user_id})
    completed_todos = mongo.db.todos.count_documents({'user_id': user_id, 'completed': True})
    active_todos = total_todos - completed_todos
    
    # Time stats
    time_pipeline = [
        {'$match': {'user_id': user_id}},
        {'$group': {
            '_id': None,
            'total_time': {'$sum': '$time_spent'},
            'avg_time': {'$avg': '$time_spent'}
        }}
    ]
    time_stats = list(mongo.db.todos.aggregate(time_pipeline))
    total_time = time_stats[0]['total_time'] if time_stats else 0
    avg_time = time_stats[0]['avg_time'] if time_stats else 0
    
    # Completion rate
    completion_rate = (completed_todos / total_todos * 100) if total_todos > 0 else 0
    
    # Category breakdown
    category_pipeline = [
        {'$match': {'user_id': user_id}},
        {'$group': {
            '_id': '$category',
            'total': {'$sum': 1},
            'completed': {'$sum': {'$cond': ['$completed', 1, 0]}}
        }}
    ]
    category_stats = list(mongo.db.todos.aggregate(category_pipeline))
    
    # Priority breakdown
    priority_pipeline = [
        {'$match': {'user_id': user_id}},
        {'$group': {
            '_id': '$priority',
            'total': {'$sum': 1},
            'completed': {'$sum': {'$cond': ['$completed', 1, 0]}}
        }}
    ]
    priority_stats = list(mongo.db.todos.aggregate(priority_pipeline))
    
    # Daily completion trend (last 7 days)
    seven_days_ago = datetime.utcnow() - timedelta(days=7)
    daily_pipeline = [
        {'$match': {
            'user_id': user_id,
            'completed': True,
            'completed_at': {'$gte': seven_days_ago}
        }},
        {'$group': {
            '_id': {
                'year': {'$year': '$completed_at'},
                'month': {'$month': '$completed_at'},
                'day': {'$dayOfMonth': '$completed_at'}
            },
            'count': {'$sum': 1}
        }},
        {'$sort': {'_id': 1}}
    ]
    daily_completions = list(mongo.db.todos.aggregate(daily_pipeline))
    
    return jsonify({
        'total_todos': total_todos,
        'completed_todos': completed_todos,
        'active_todos': active_todos,
        'completion_rate': round(completion_rate, 1),
        'total_time_minutes': int(total_time),
        'avg_time_minutes': round(avg_time, 1),
        'category_breakdown': category_stats,
        'priority_breakdown': priority_stats,
        'daily_completions': daily_completions
    }), 200

# AI Task Suggestions with Multiple Providers
@app.route('/api/ai/suggestions', methods=['POST'])
@token_required
def get_ai_suggestions(current_user):
    data = request.get_json()
    context = data.get('context', '')
    
    # Get user's recent tasks for context
    recent_todos = list(mongo.db.todos.find(
        {'user_id': str(current_user['_id'])}
    ).sort('created_at', -1).limit(10))
    
    # Prepare context for AI
    task_context = []
    for todo in recent_todos:
        task_context.append({
            'text': todo['text'],
            'category': todo['category'],
            'priority': todo['priority'],
            'completed': todo['completed']
        })
    
    # Get AI provider settings
    api_settings = current_user.get('api_settings', {})
    ai_provider = api_settings.get('ai_provider', 'gemini')
    
    # Get appropriate API key based on provider
    if ai_provider == 'gemini':
        api_key = api_settings.get('gemini_api_key', '') or os.getenv('GEMINI_API_KEY')
        if not api_key:
            # Smart fallback suggestions based on user's actual tasks
            incomplete_tasks = [t for t in recent_todos if not t['completed']]
            total_tasks = len(recent_todos)
            
            fallback_suggestions = [
                {
                    "text": "Set up your AI API key in Profile Settings for personalized suggestions",
                    "category": "work",
                    "priority": "medium",
                    "reasoning": "AI can provide smarter task recommendations based on your patterns",
                    "action_type": "optimize"
                }
            ]
            
            if incomplete_tasks:
                fallback_suggestions.append({
                    "text": f"Focus on completing your {len(incomplete_tasks)} pending tasks",
                    "category": "work",
                    "priority": "high",
                    "reasoning": "Completing existing tasks before adding new ones improves productivity",
                    "action_type": "focus"
                })
                
                # Suggest specific incomplete task
                if incomplete_tasks:
                    fallback_suggestions.append({
                        "text": f"Work on: {incomplete_tasks[0]['text'][:50]}{'...' if len(incomplete_tasks[0]['text']) > 50 else ''}",
                        "category": incomplete_tasks[0]['category'],
                        "priority": incomplete_tasks[0]['priority'],
                        "reasoning": "This task is waiting for your attention",
                        "action_type": "focus"
                    })
            else:
                fallback_suggestions.append({
                    "text": "Plan your next goals and create actionable tasks",
                    "category": "personal",
                    "priority": "medium",
                    "reasoning": "Good planning leads to better productivity",
                    "action_type": "plan"
                })
            return jsonify({
                'suggestions': fallback_suggestions,
                'message': 'Gemini API key not configured. Please set it in your profile settings for AI-powered suggestions.',
                'fallback': True,
                'setup_required': True
            }), 200
            
    elif ai_provider == 'openai':
        api_key = api_settings.get('openai_api_key', '') or os.getenv('OPENAI_API_KEY')
        if not api_key:
            return jsonify({'message': 'OpenAI API key not configured. Please set it in your profile settings.'}), 500
    elif ai_provider == 'claude':
        api_key = api_settings.get('claude_api_key', '') or os.getenv('CLAUDE_API_KEY')
        if not api_key:
            return jsonify({'message': 'Claude API key not configured. Please set it in your profile settings.'}), 500
    elif ai_provider == 'custom':
        api_key = api_settings.get('custom_api_key', '')
        custom_endpoint = api_settings.get('custom_api_endpoint', '')
        if not api_key or not custom_endpoint:
            return jsonify({'message': 'Custom API settings not configured. Please set API key and endpoint in your profile settings.'}), 500
    else:
        return jsonify({'message': 'Invalid AI provider selected.'}), 500
    
    # Analyze user's task patterns
    total_tasks = len(recent_todos)
    completed_tasks = len([t for t in recent_todos if t['completed']])
    pending_tasks = total_tasks - completed_tasks
    
    # Category analysis
    categories = {}
    priorities = {}
    for todo in recent_todos:
        cat = todo['category']
        pri = todo['priority']
        categories[cat] = categories.get(cat, 0) + 1
        priorities[pri] = priorities.get(pri, 0) + 1
    
    most_used_category = max(categories.keys(), default='work') if categories else 'work'
    most_used_priority = max(priorities.keys(), default='medium') if priorities else 'medium'
    
    # Get incomplete tasks for analysis
    incomplete_tasks = [t for t in recent_todos if not t['completed']]
    
    prompt = f"""
    You are an intelligent task management assistant. Analyze the user's current tasks and provide smart, actionable suggestions.
    
    USER'S CURRENT SITUATION:
    - Total tasks: {total_tasks}
    - Completed: {completed_tasks}
    - Pending: {pending_tasks}
    - Most used category: {most_used_category}
    - Most used priority: {most_used_priority}
    - User context: {context}
    
    CURRENT INCOMPLETE TASKS:
    {json.dumps([t['text'] for t in incomplete_tasks[:5]])}
    
    RECENT TASK PATTERNS:
    {json.dumps(task_context)}
    
    Based on this analysis, provide 3-5 SMART suggestions that:
    1. Help complete or organize existing tasks (not just add new ones)
    2. Address productivity gaps or patterns you notice
    3. Suggest next steps for incomplete tasks
    4. Recommend task management improvements
    5. Are contextually relevant to their work style
    
    EXAMPLES OF SMART SUGGESTIONS:
    - "Break down '[specific incomplete task]' into smaller steps"
    - "Set a 25-minute timer to focus on '[high priority incomplete task]'"
    - "Review and prioritize your {pending_tasks} pending tasks"
    - "Schedule time blocks for your {most_used_category} tasks"
    - "Archive completed tasks to declutter your list"
    
    Return response in JSON format:
    {{
        "suggestions": [
            {{
                "text": "Specific actionable suggestion based on their actual tasks",
                "category": "work|personal|shopping|health",
                "priority": "low|medium|high",
                "reasoning": "Why this suggestion helps their current situation",
                "action_type": "organize|focus|plan|review|optimize"
            }}
        ]
    }}
    """
    
    try:
        # Make API call based on provider
        if ai_provider == 'gemini':
            url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
            payload = {
                "contents": [{
                    "parts": [{
                        "text": prompt
                    }]
                }]
            }
            response = requests.post(url, json=payload)
            
        elif ai_provider == 'openai':
            url = "https://api.openai.com/v1/chat/completions"
            headers = {
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json"
            }
            payload = {
                "model": "gpt-3.5-turbo",
                "messages": [{"role": "user", "content": prompt}],
                "max_tokens": 1000
            }
            response = requests.post(url, json=payload, headers=headers)
            
        elif ai_provider == 'claude':
            url = "https://api.anthropic.com/v1/messages"
            headers = {
                "x-api-key": api_key,
                "Content-Type": "application/json",
                "anthropic-version": "2023-06-01"
            }
            payload = {
                "model": "claude-3-sonnet-20240229",
                "max_tokens": 1000,
                "messages": [{"role": "user", "content": prompt}]
            }
            response = requests.post(url, json=payload, headers=headers)
            
        elif ai_provider == 'custom':
            headers = {
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json"
            }
            payload = {
                "prompt": prompt,
                "max_tokens": 1000
            }
            response = requests.post(custom_endpoint, json=payload, headers=headers)
        
        if response.status_code == 200:
            result = response.json()
            
            # Extract AI response based on provider
            if ai_provider == 'gemini':
                ai_response = result['candidates'][0]['content']['parts'][0]['text']
            elif ai_provider == 'openai':
                ai_response = result['choices'][0]['message']['content']
            elif ai_provider == 'claude':
                ai_response = result['content'][0]['text']
            elif ai_provider == 'custom':
                ai_response = result.get('response', result.get('text', str(result)))
            
            # Try to parse JSON from AI response
            try:
                # Clean the response (remove markdown formatting if present)
                clean_response = ai_response.strip()
                if clean_response.startswith('```json'):
                    clean_response = clean_response[7:]
                if clean_response.endswith('```'):
                    clean_response = clean_response[:-3]
                
                suggestions = json.loads(clean_response)
                return jsonify(suggestions), 200
            except json.JSONDecodeError:
                # Fallback: return raw response
                return jsonify({
                    'suggestions': [],
                    'raw_response': ai_response,
                    'message': 'AI response received but could not parse JSON',
                    'provider': ai_provider
                }), 200
        else:
            # Log the error response for debugging
            error_details = {
                'status_code': response.status_code,
                'response_text': response.text[:500],  # First 500 chars
                'provider': ai_provider,
                'url': response.url if hasattr(response, 'url') else 'unknown'
            }
            
            # Check for specific error types
            if response.status_code == 404:
                error_message = f'API endpoint not found for {ai_provider}. Please check your API key and provider settings.'
            elif response.status_code == 401:
                error_message = f'Invalid API key for {ai_provider}. Please check your API key in profile settings.'
            elif response.status_code == 403:
                error_message = f'Access forbidden for {ai_provider}. Please check your API key permissions.'
            elif response.status_code == 429:
                error_message = f'Rate limit exceeded for {ai_provider}. Please try again later.'
            else:
                error_message = f'Failed to get AI suggestions from {ai_provider}. Status: {response.status_code}'
            
            return jsonify({
                'message': error_message,
                'error_details': error_details,
                'fallback_available': True
            }), 500
            
    except Exception as e:
        # Log the full error for debugging
        print(f"AI Suggestions Error: {str(e)}")
        print(f"Provider: {ai_provider}")
        print(f"API Key Length: {len(api_key) if api_key else 0}")
        
        # Smart fallback suggestions based on user's actual tasks
        incomplete_tasks = [t for t in recent_todos if not t['completed']]
        total_tasks = len(recent_todos)
        
        fallback_suggestions = []
        
        if incomplete_tasks:
            # Suggest focusing on incomplete tasks
            high_priority_incomplete = [t for t in incomplete_tasks if t['priority'] == 'high']
            if high_priority_incomplete:
                fallback_suggestions.append({
                    "text": f"Focus on completing: {high_priority_incomplete[0]['text'][:50]}...",
                    "category": high_priority_incomplete[0]['category'],
                    "priority": "high",
                    "reasoning": "You have high-priority tasks that need attention",
                    "action_type": "focus"
                })
            
            fallback_suggestions.append({
                "text": f"Review your {len(incomplete_tasks)} pending tasks and prioritize them",
                "category": "work",
                "priority": "medium",
                "reasoning": "Organization helps improve productivity",
                "action_type": "organize"
            })
        
        if total_tasks > 10:
            fallback_suggestions.append({
                "text": "Archive completed tasks to declutter your task list",
                "category": "work",
                "priority": "low",
                "reasoning": "A clean task list improves focus",
                "action_type": "organize"
            })
        
        # Always include a wellness suggestion
        fallback_suggestions.append({
            "text": "Take a 10-minute break to recharge and refocus",
            "category": "health",
            "priority": "low",
            "reasoning": "Regular breaks improve productivity and well-being",
            "action_type": "optimize"
        })
        
        # If no tasks, suggest getting started
        if total_tasks == 0:
            fallback_suggestions = [
                {
                    "text": "Create your first task to get started with productivity tracking",
                    "category": "work",
                    "priority": "medium",
                    "reasoning": "Starting with one task builds momentum",
                    "action_type": "plan"
                },
                {
                    "text": "Set up your daily routine and goals",
                    "category": "personal",
                    "priority": "medium",
                    "reasoning": "Planning helps establish good habits",
                    "action_type": "plan"
                }
            ]
        
        return jsonify({
            'suggestions': fallback_suggestions,
            'message': f'AI service temporarily unavailable. Showing fallback suggestions.',
            'error_details': str(e),
            'provider': ai_provider,
            'fallback': True
        }), 200

# Advanced AI Task Analysis & Insights
@app.route('/api/ai/analyze', methods=['POST'])
@token_required
def analyze_task_with_ai(current_user):
    data = request.get_json()
    task_text = data.get('task_text', '')
    
    if not task_text:
        return jsonify({'message': 'Task text required'}), 400
    
    # Get user's complete task history for context
    all_todos = list(mongo.db.todos.find({'user_id': str(current_user['_id'])}).sort('created_at', -1))
    completed_todos = [t for t in all_todos if t.get('completed')]
    
    # Calculate user patterns
    avg_completion_time = 0
    if completed_todos:
        total_time = sum([t.get('time_spent', 0) for t in completed_todos])
        avg_completion_time = total_time / len(completed_todos) if completed_todos else 0
    
    # Category performance analysis
    category_stats = {}
    for todo in all_todos:
        cat = todo['category']
        if cat not in category_stats:
            category_stats[cat] = {'total': 0, 'completed': 0, 'avg_time': 0}
        category_stats[cat]['total'] += 1
        if todo.get('completed'):
            category_stats[cat]['completed'] += 1
            category_stats[cat]['avg_time'] += todo.get('time_spent', 0)
    
    # Calculate completion rates by category
    for cat in category_stats:
        if category_stats[cat]['completed'] > 0:
            category_stats[cat]['avg_time'] /= category_stats[cat]['completed']
            category_stats[cat]['completion_rate'] = (category_stats[cat]['completed'] / category_stats[cat]['total']) * 100
    
    # Get AI provider settings
    api_settings = current_user.get('api_settings', {})
    ai_provider = api_settings.get('ai_provider', 'gemini')
    
    # Get appropriate API key based on provider
    if ai_provider == 'gemini':
        api_key = api_settings.get('gemini_api_key', '') or os.getenv('GEMINI_API_KEY')
    elif ai_provider == 'openai':
        api_key = api_settings.get('openai_api_key', '') or os.getenv('OPENAI_API_KEY')
    elif ai_provider == 'claude':
        api_key = api_settings.get('claude_api_key', '') or os.getenv('CLAUDE_API_KEY')
    elif ai_provider == 'custom':
        api_key = api_settings.get('custom_api_key', '')
        custom_endpoint = api_settings.get('custom_api_endpoint', '')
    else:
        return jsonify({'message': 'Invalid AI provider selected.'}), 500
    
    if not api_key:
        return jsonify({'message': f'{ai_provider.title()} API key not configured. Please set it in your profile settings.'}), 500
    
    prompt = f"""
    You are an expert productivity consultant with deep knowledge of task management, psychology, and efficiency optimization. 
    Analyze this task with the context of the user's historical performance and provide comprehensive insights.
    
    TASK TO ANALYZE: "{task_text}"
    
    USER'S HISTORICAL CONTEXT:
    - Total tasks completed: {len(completed_todos)}
    - Average completion time: {avg_completion_time:.1f} minutes
    - Category performance: {json.dumps(category_stats)}
    - Most productive category: {max(category_stats.keys(), key=lambda x: category_stats[x].get('completion_rate', 0)) if category_stats else 'work'}
    
    PROVIDE COMPREHENSIVE ANALYSIS:
    
    1. TASK CLASSIFICATION:
       - Optimal category based on user's success patterns
       - Priority level considering user's workload and patterns
       - Complexity assessment (simple/moderate/complex)
       - Task type (creative/analytical/administrative/physical)
    
    2. TIME & EFFORT ESTIMATION:
       - Realistic time estimate based on user's historical performance
       - Energy level required (low/medium/high)
       - Best time of day to tackle this task
       - Potential obstacles and time buffers
    
    3. STRATEGIC BREAKDOWN:
       - Break complex tasks into actionable subtasks
       - Identify dependencies and prerequisites
       - Suggest optimal sequencing
       - Highlight critical path items
    
    4. OPTIMIZATION STRATEGIES:
       - Specific techniques for this task type
       - Tools and resources that could help
       - Ways to batch similar tasks
       - Automation opportunities
    
    5. PSYCHOLOGICAL INSIGHTS:
       - Motivation techniques for this specific task
       - Ways to overcome procrastination
       - Reward systems that would work
       - Accountability measures
    
    6. CONTEXTUAL RECOMMENDATIONS:
       - Best environment/setting for this task
       - Preparation steps needed
       - Follow-up actions required
       - Success metrics to track
    
    Return response in JSON format:
    {{
        "category": "optimal category based on user patterns",
        "priority": "low|medium|high",
        "complexity": "simple|moderate|complex",
        "task_type": "creative|analytical|administrative|physical|mixed",
        "estimated_time": number_in_minutes,
        "energy_required": "low|medium|high",
        "best_time": "morning|afternoon|evening|flexible",
        "subtasks": ["detailed actionable subtasks"],
        "prerequisites": ["things needed before starting"],
        "tools_needed": ["helpful tools or resources"],
        "optimization_tips": ["specific efficiency strategies"],
        "motivation_techniques": ["psychological approaches"],
        "success_metrics": ["how to measure completion"],
        "potential_obstacles": ["likely challenges and solutions"],
        "batch_opportunities": ["similar tasks to group together"],
        "follow_up_actions": ["what to do after completion"],
        "reasoning": "comprehensive explanation of all recommendations",
        "confidence_score": number_0_to_100,
        "user_success_prediction": "likelihood of completion based on patterns"
    }}
    """
    
    try:
        # Make API call based on provider
        if ai_provider == 'gemini':
            url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
            payload = {
                "contents": [{
                    "parts": [{
                        "text": prompt
                    }]
                }]
            }
            response = requests.post(url, json=payload)
            
        elif ai_provider == 'openai':
            url = "https://api.openai.com/v1/chat/completions"
            headers = {
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json"
            }
            payload = {
                "model": "gpt-3.5-turbo",
                "messages": [{"role": "user", "content": prompt}],
                "max_tokens": 1000
            }
            response = requests.post(url, json=payload, headers=headers)
            
        elif ai_provider == 'claude':
            url = "https://api.anthropic.com/v1/messages"
            headers = {
                "x-api-key": api_key,
                "Content-Type": "application/json",
                "anthropic-version": "2023-06-01"
            }
            payload = {
                "model": "claude-3-sonnet-20240229",
                "max_tokens": 1000,
                "messages": [{"role": "user", "content": prompt}]
            }
            response = requests.post(url, json=payload, headers=headers)
            
        elif ai_provider == 'custom':
            headers = {
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json"
            }
            payload = {
                "prompt": prompt,
                "max_tokens": 1000
            }
            response = requests.post(custom_endpoint, json=payload, headers=headers)
        
        if response.status_code == 200:
            result = response.json()
            
            # Extract AI response based on provider
            if ai_provider == 'gemini':
                ai_response = result['candidates'][0]['content']['parts'][0]['text']
            elif ai_provider == 'openai':
                ai_response = result['choices'][0]['message']['content']
            elif ai_provider == 'claude':
                ai_response = result['content'][0]['text']
            elif ai_provider == 'custom':
                ai_response = result.get('response', result.get('text', str(result)))
            
            try:
                # Clean the response
                clean_response = ai_response.strip()
                if clean_response.startswith('```json'):
                    clean_response = clean_response[7:]
                if clean_response.endswith('```'):
                    clean_response = clean_response[:-3]
                
                analysis = json.loads(clean_response)
                return jsonify(analysis), 200
            except json.JSONDecodeError:
                return jsonify({
                    'raw_response': ai_response,
                    'message': 'AI response received but could not parse JSON',
                    'provider': ai_provider
                }), 200
        else:
            return jsonify({'message': f'Failed to analyze task with {ai_provider}. Status: {response.status_code}'}), 500
            
    except Exception as e:
        return jsonify({'message': f'AI service error: {str(e)}', 'provider': ai_provider}), 500

# AI-Powered Daily Planning
@app.route('/api/ai/plan-day', methods=['POST'])
@token_required
def ai_plan_day(current_user):
    data = request.get_json()
    available_hours = data.get('available_hours', 8)
    energy_level = data.get('energy_level', 'medium')  # low, medium, high
    focus_areas = data.get('focus_areas', [])  # work, personal, health, etc.
    
    # Get user's tasks and patterns
    all_todos = list(mongo.db.todos.find({'user_id': str(current_user['_id'])}).sort('created_at', -1))
    incomplete_todos = [t for t in all_todos if not t.get('completed')]
    
    # Get user's productivity patterns from activities
    activities = list(mongo.db.activities.find({'user_id': str(current_user['_id'])}).sort('created_at', -1).limit(50))
    
    # Analyze when user is most productive
    productivity_hours = {}
    for activity in activities:
        if activity.get('type') == 'task_completed':
            hour = activity['created_at'].hour
            productivity_hours[hour] = productivity_hours.get(hour, 0) + 1
    
    api_settings = current_user.get('api_settings', {})
    ai_provider = api_settings.get('ai_provider', 'gemini')
    api_key = api_settings.get('gemini_api_key', '') or os.getenv('GEMINI_API_KEY')
    
    if not api_key:
        return jsonify({'message': 'AI API key required for daily planning'}), 400
    
    prompt = f"""
    You are an expert productivity coach and time management consultant. Create an optimal daily plan for this user.
    
    USER'S SITUATION:
    - Available hours today: {available_hours}
    - Current energy level: {energy_level}
    - Focus areas: {focus_areas}
    - Incomplete tasks: {len(incomplete_todos)}
    - Most productive hours: {sorted(productivity_hours.items(), key=lambda x: x[1], reverse=True)[:3]}
    
    INCOMPLETE TASKS TO CONSIDER:
    {json.dumps([{
        'text': t['text'][:100],
        'category': t['category'],
        'priority': t['priority'],
        'estimated_time': t.get('estimated_time', 30)
    } for t in incomplete_todos[:10]])}
    
    CREATE A COMPREHENSIVE DAILY PLAN:
    
    1. PRIORITIZED TASK SCHEDULE:
       - Rank tasks by importance and energy requirements
       - Assign optimal time slots based on user's productivity patterns
       - Include realistic time estimates and buffers
       - Balance different types of work
    
    2. ENERGY MANAGEMENT:
       - Match high-energy tasks with peak productivity hours
       - Schedule breaks and recovery periods
       - Include energy-boosting activities
       - Plan for energy dips
    
    3. FOCUS OPTIMIZATION:
       - Group similar tasks together
       - Minimize context switching
       - Include deep work blocks
       - Plan for shallow work periods
    
    4. WELLNESS INTEGRATION:
       - Include movement and exercise
       - Schedule proper meal breaks
       - Add mindfulness moments
       - Ensure work-life balance
    
    Return response in JSON format:
    {{
        "daily_schedule": [
            {{
                "time_slot": "09:00-10:30",
                "activity": "specific task or activity",
                "category": "work|personal|health|break",
                "energy_required": "low|medium|high",
                "focus_type": "deep|shallow|creative|administrative",
                "estimated_duration": minutes,
                "reasoning": "why this time slot is optimal"
            }}
        ],
        "priority_tasks": ["top 3 must-do tasks for today"],
        "energy_strategy": "how to manage energy throughout the day",
        "focus_blocks": ["dedicated focus periods and their purposes"],
        "break_schedule": ["when and what type of breaks to take"],
        "success_tips": ["specific advice for making this plan work"],
        "contingency_plan": "what to do if things don't go as planned",
        "evening_review": ["questions to ask yourself at end of day"],
        "total_productive_hours": number,
        "wellness_score": number_0_to_100
    }}
    """
    
    try:
        if ai_provider == 'gemini':
            url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
            payload = {
                "contents": [{
                    "parts": [{
                        "text": prompt
                    }]
                }]
            }
            response = requests.post(url, json=payload)
            
            if response.status_code == 200:
                result = response.json()
                ai_response = result['candidates'][0]['content']['parts'][0]['text']
                
                try:
                    clean_response = ai_response.strip()
                    if clean_response.startswith('```json'):
                        clean_response = clean_response[7:]
                    if clean_response.endswith('```'):
                        clean_response = clean_response[:-3]
                    
                    plan = json.loads(clean_response)
                    
                    # Log the planning activity
                    mongo.db.activities.insert_one({
                        'user_id': str(current_user['_id']),
                        'type': 'ai_daily_plan',
                        'description': f'Generated AI daily plan for {available_hours} hours',
                        'created_at': datetime.utcnow()
                    })
                    
                    return jsonify(plan), 200
                except json.JSONDecodeError:
                    return jsonify({
                        'raw_response': ai_response,
                        'message': 'AI response received but could not parse JSON'
                    }), 200
            else:
                return jsonify({'message': f'Failed to generate daily plan. Status: {response.status_code}'}), 500
                
    except Exception as e:
        return jsonify({'message': f'AI service error: {str(e)}'}), 500

# AI Task Optimization & Workflow Analysis
@app.route('/api/ai/optimize-workflow', methods=['POST'])
@token_required
def ai_optimize_workflow(current_user):
    # Get comprehensive user data
    all_todos = list(mongo.db.todos.find({'user_id': str(current_user['_id'])}).sort('created_at', -1))
    activities = list(mongo.db.activities.find({'user_id': str(current_user['_id'])}).sort('created_at', -1).limit(100))
    
    # Analyze patterns
    completion_patterns = {}
    time_patterns = {}
    category_efficiency = {}
    
    for todo in all_todos:
        if todo.get('completed'):
            # Day of week patterns
            day = todo.get('completed_at', todo['created_at']).weekday()
            completion_patterns[day] = completion_patterns.get(day, 0) + 1
            
            # Time spent patterns
            time_spent = todo.get('time_spent', 0)
            category = todo['category']
            if category not in category_efficiency:
                category_efficiency[category] = {'total_time': 0, 'count': 0, 'avg_time': 0}
            category_efficiency[category]['total_time'] += time_spent
            category_efficiency[category]['count'] += 1
    
    # Calculate averages
    for cat in category_efficiency:
        if category_efficiency[cat]['count'] > 0:
            category_efficiency[cat]['avg_time'] = category_efficiency[cat]['total_time'] / category_efficiency[cat]['count']
    
    api_settings = current_user.get('api_settings', {})
    api_key = api_settings.get('gemini_api_key', '') or os.getenv('GEMINI_API_KEY')
    
    if not api_key:
        return jsonify({'message': 'AI API key required for workflow optimization'}), 400
    
    prompt = f"""
    You are a world-class productivity expert and workflow optimization specialist. Analyze this user's complete task management patterns and provide comprehensive optimization recommendations.
    
    USER'S COMPLETE PERFORMANCE DATA:
    - Total tasks: {len(all_todos)}
    - Completed tasks: {len([t for t in all_todos if t.get('completed')])}
    - Completion rate: {(len([t for t in all_todos if t.get('completed')]) / len(all_todos) * 100) if all_todos else 0:.1f}%
    - Most productive days: {sorted(completion_patterns.items(), key=lambda x: x[1], reverse=True)}
    - Category efficiency: {json.dumps(category_efficiency)}
    - Recent activities: {len(activities)} logged actions
    
    CURRENT CHALLENGES:
    - Incomplete tasks: {len([t for t in all_todos if not t.get('completed')])}
    - Average time per task: {sum([t.get('time_spent', 0) for t in all_todos if t.get('completed')]) / len([t for t in all_todos if t.get('completed')]) if [t for t in all_todos if t.get('completed')] else 0:.1f} minutes
    
    PROVIDE COMPREHENSIVE WORKFLOW OPTIMIZATION:
    
    1. PRODUCTIVITY ANALYSIS:
       - Identify peak performance patterns
       - Highlight efficiency bottlenecks
       - Analyze task completion trends
       - Spot procrastination patterns
    
    2. WORKFLOW OPTIMIZATION:
       - Recommend optimal task sequencing
       - Suggest batching strategies
       - Identify automation opportunities
       - Propose time-blocking methods
    
    3. CATEGORY OPTIMIZATION:
       - Optimize each task category approach
       - Suggest category-specific strategies
       - Recommend focus techniques per category
       - Identify cross-category synergies
    
    4. TIME MANAGEMENT IMPROVEMENTS:
       - Optimize time estimation accuracy
       - Suggest buffer time strategies
       - Recommend break patterns
       - Propose deadline management
    
    5. HABIT FORMATION:
       - Identify keystone habits to develop
       - Suggest habit stacking opportunities
       - Recommend consistency strategies
       - Propose accountability systems
    
    6. TECHNOLOGY INTEGRATION:
       - Suggest productivity tools
       - Recommend automation setups
       - Propose tracking improvements
       - Identify integration opportunities
    
    Return response in JSON format:
    {{
        "productivity_score": number_0_to_100,
        "key_insights": ["major patterns and findings"],
        "optimization_opportunities": [
            {{
                "area": "specific area to improve",
                "current_state": "what's happening now",
                "recommended_change": "specific action to take",
                "expected_impact": "predicted improvement",
                "implementation_difficulty": "easy|medium|hard",
                "time_to_see_results": "timeframe for improvement"
            }}
        ],
        "workflow_recommendations": [
            {{
                "strategy": "specific workflow strategy",
                "description": "how to implement it",
                "benefits": ["expected benefits"],
                "tools_needed": ["required tools or resources"],
                "success_metrics": ["how to measure success"]
            }}
        ],
        "daily_routine_suggestions": {{
            "morning_routine": ["optimal morning tasks"],
            "peak_hours_usage": "how to use most productive hours",
            "afternoon_strategy": "approach for afternoon productivity",
            "evening_routine": ["end-of-day optimization"]
        }},
        "habit_recommendations": [
            {{
                "habit": "specific habit to develop",
                "trigger": "when to do it",
                "reward": "how to reinforce it",
                "tracking": "how to measure progress"
            }}
        ],
        "technology_stack": {{
            "recommended_tools": ["productivity tools to consider"],
            "automation_ideas": ["processes to automate"],
            "integration_opportunities": ["ways to connect systems"]
        }},
        "next_steps": ["immediate actions to take"],
        "long_term_goals": ["bigger improvements to work toward"],
        "success_probability": number_0_to_100
    }}
    """
    
    try:
        url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
        payload = {
            "contents": [{
                "parts": [{
                    "text": prompt
                }]
            }]
        }
        response = requests.post(url, json=payload)
        
        if response.status_code == 200:
            result = response.json()
            ai_response = result['candidates'][0]['content']['parts'][0]['text']
            
            try:
                clean_response = ai_response.strip()
                if clean_response.startswith('```json'):
                    clean_response = clean_response[7:]
                if clean_response.endswith('```'):
                    clean_response = clean_response[:-3]
                
                optimization = json.loads(clean_response)
                
                # Log the optimization activity
                mongo.db.activities.insert_one({
                    'user_id': str(current_user['_id']),
                    'type': 'ai_workflow_optimization',
                    'description': 'Generated comprehensive workflow optimization analysis',
                    'created_at': datetime.utcnow()
                })
                
                return jsonify(optimization), 200
            except json.JSONDecodeError:
                return jsonify({
                    'raw_response': ai_response,
                    'message': 'AI response received but could not parse JSON'
                }), 200
        else:
            return jsonify({'message': f'Failed to generate optimization. Status: {response.status_code}'}), 500
            
    except Exception as e:
        return jsonify({'message': f'AI service error: {str(e)}'}), 500

# AI Context-Aware Smart Suggestions
@app.route('/api/ai/smart-suggestions', methods=['POST'])
@token_required
def ai_smart_suggestions(current_user):
    data = request.get_json()
    context_type = data.get('context_type', 'general')  # general, morning, afternoon, evening, weekend
    current_mood = data.get('mood', 'neutral')  # energetic, tired, stressed, focused, creative
    available_time = data.get('available_time', 30)  # minutes
    
    # Get comprehensive user context
    all_todos = list(mongo.db.todos.find({'user_id': str(current_user['_id'])}).sort('created_at', -1))
    recent_activities = list(mongo.db.activities.find({'user_id': str(current_user['_id'])}).sort('created_at', -1).limit(20))
    
    # Analyze current situation
    incomplete_todos = [t for t in all_todos if not t.get('completed')]
    overdue_todos = []
    high_priority_todos = [t for t in incomplete_todos if t.get('priority') == 'high']
    
    # Check for patterns in recent activities
    recent_completions = [a for a in recent_activities if a.get('type') == 'task_completed']
    recent_categories = [a.get('description', '').split(':')[-1].strip() for a in recent_completions[:5]]
    
    api_settings = current_user.get('api_settings', {})
    api_key = api_settings.get('gemini_api_key', '') or os.getenv('GEMINI_API_KEY')
    
    if not api_key:
        # Smart fallback based on context
        smart_fallbacks = []
        
        if context_type == 'morning':
            smart_fallbacks = [
                {
                    "text": "Review your top 3 priorities for today",
                    "category": "work",
                    "priority": "high",
                    "reasoning": "Morning is ideal for planning and prioritization",
                    "action_type": "plan",
                    "estimated_time": 10
                },
                {
                    "text": "Tackle your most challenging task while energy is high",
                    "category": "work",
                    "priority": "high",
                    "reasoning": "Morning energy is best used for difficult tasks",
                    "action_type": "focus",
                    "estimated_time": available_time
                }
            ]
        elif context_type == 'afternoon':
            smart_fallbacks = [
                {
                    "text": "Handle administrative tasks and emails",
                    "category": "work",
                    "priority": "medium",
                    "reasoning": "Afternoon is good for routine administrative work",
                    "action_type": "organize",
                    "estimated_time": available_time
                }
            ]
        elif context_type == 'evening':
            smart_fallbacks = [
                {
                    "text": "Review today's accomplishments and plan tomorrow",
                    "category": "personal",
                    "priority": "medium",
                    "reasoning": "Evening reflection improves next-day productivity",
                    "action_type": "review",
                    "estimated_time": 15
                }
            ]
        
        if high_priority_todos:
            smart_fallbacks.insert(0, {
                "text": f"Focus on high-priority task: {high_priority_todos[0]['text'][:50]}...",
                "category": high_priority_todos[0]['category'],
                "priority": "high",
                "reasoning": "You have high-priority tasks waiting",
                "action_type": "focus",
                "estimated_time": available_time
            })
        
        return jsonify({
            'suggestions': smart_fallbacks[:3],
            'context_analysis': {
                'available_time': available_time,
                'context_type': context_type,
                'high_priority_count': len(high_priority_todos),
                'total_incomplete': len(incomplete_todos)
            },
            'message': 'Smart fallback suggestions based on context',
            'fallback': True
        }), 200
    
    prompt = f"""
    You are an advanced AI productivity assistant with deep understanding of human psychology, circadian rhythms, and optimal performance patterns. 
    Provide highly contextual and personalized suggestions based on the user's current situation.
    
    CURRENT CONTEXT:
    - Time context: {context_type}
    - User's mood/energy: {current_mood}
    - Available time: {available_time} minutes
    - Incomplete tasks: {len(incomplete_todos)}
    - High priority tasks: {len(high_priority_todos)}
    - Recent activity pattern: {recent_categories}
    
    USER'S TASK LANDSCAPE:
    High Priority Tasks: {json.dumps([t['text'][:50] for t in high_priority_todos[:3]])}
    Recent Categories Worked On: {recent_categories}
    
    PROVIDE CONTEXT-AWARE SUGGESTIONS:
    
    Consider these factors:
    1. CIRCADIAN OPTIMIZATION:
       - Morning: Complex cognitive tasks, planning, creative work
       - Afternoon: Collaborative work, meetings, routine tasks
       - Evening: Reflection, planning, light administrative work
    
    2. MOOD-ENERGY MATCHING:
       - Energetic: Challenging tasks, new projects, problem-solving
       - Tired: Routine tasks, organization, easy completions
       - Stressed: Calming activities, small wins, organization
       - Focused: Deep work, complex analysis, important projects
       - Creative: Brainstorming, design, innovative thinking
    
    3. TIME OPTIMIZATION:
       - Short time (5-15 min): Quick wins, organization, planning
       - Medium time (15-45 min): Focused work sessions, specific tasks
       - Long time (45+ min): Deep work, complex projects, learning
    
    4. PSYCHOLOGICAL FACTORS:
       - Build momentum with quick wins
       - Match task difficulty to current capacity
       - Consider context switching costs
       - Optimize for flow state potential
    
    Return response in JSON format:
    {{
        "suggestions": [
            {{
                "text": "specific contextual suggestion",
                "category": "work|personal|shopping|health",
                "priority": "low|medium|high",
                "reasoning": "why this is perfect for current context",
                "action_type": "focus|organize|plan|review|create|learn",
                "estimated_time": minutes,
                "energy_match": "how well this matches current energy",
                "flow_potential": "likelihood of entering flow state",
                "momentum_builder": "whether this builds positive momentum",
                "context_optimization": "how this leverages current context"
            }}
        ],
        "context_analysis": {{
            "optimal_activities": ["best activities for current context"],
            "energy_assessment": "analysis of current energy state",
            "time_utilization": "how to best use available time",
            "momentum_strategy": "approach to build positive momentum"
        }},
        "next_context_preparation": {{
            "upcoming_context": "what context comes next",
            "preparation_suggestions": ["how to prepare for next phase"],
            "transition_activities": ["activities to bridge contexts"]
        }},
        "personalization_insights": {{
            "pattern_recognition": ["patterns noticed in user behavior"],
            "optimization_opportunities": ["ways to improve based on patterns"],
            "success_predictors": ["factors that predict user success"]
        }}
    }}
    """
    
    try:
        url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
        payload = {
            "contents": [{
                "parts": [{
                    "text": prompt
                }]
            }]
        }
        response = requests.post(url, json=payload)
        
        if response.status_code == 200:
            result = response.json()
            ai_response = result['candidates'][0]['content']['parts'][0]['text']
            
            try:
                clean_response = ai_response.strip()
                if clean_response.startswith('```json'):
                    clean_response = clean_response[7:]
                if clean_response.endswith('```'):
                    clean_response = clean_response[:-3]
                
                suggestions = json.loads(clean_response)
                
                # Log the smart suggestion activity
                mongo.db.activities.insert_one({
                    'user_id': str(current_user['_id']),
                    'type': 'ai_smart_suggestions',
                    'description': f'Generated context-aware suggestions for {context_type} context',
                    'created_at': datetime.utcnow()
                })
                
                return jsonify(suggestions), 200
            except json.JSONDecodeError:
                return jsonify({
                    'raw_response': ai_response,
                    'message': 'AI response received but could not parse JSON'
                }), 200
        else:
            return jsonify({'message': f'Failed to generate smart suggestions. Status: {response.status_code}'}), 500
            
    except Exception as e:
        return jsonify({'message': f'AI service error: {str(e)}'}), 500

# Ticket Routes
@app.route('/api/tickets', methods=['GET'])
@token_required
def get_tickets(current_user):
    tickets = list(mongo.db.tickets.find({'user_id': str(current_user['_id'])}).sort('created_at', -1))
    for ticket in tickets:
        ticket['_id'] = str(ticket['_id'])
        ticket['created_at'] = ticket['created_at'].isoformat()
        if ticket.get('updated_at'):
            ticket['updated_at'] = ticket['updated_at'].isoformat()
    return jsonify(tickets), 200

# Get unique clients for filter
@app.route('/api/tickets/clients', methods=['GET'])
@token_required
def get_ticket_clients(current_user):
    clients = mongo.db.tickets.distinct('client_name', {'user_id': str(current_user['_id'])})
    return jsonify(clients), 200

@app.route('/api/tickets', methods=['POST'])
@token_required
def create_ticket(current_user):
    data = request.get_json()
    
    # Check if ticket_id already exists
    if data.get('ticket_id'):
        existing = mongo.db.tickets.find_one({'ticket_id': data.get('ticket_id'), 'user_id': str(current_user['_id'])})
        if existing:
            return jsonify({'message': 'Ticket ID already exists'}), 400
    
    ticket = {
        'ticket_id': data.get('ticket_id'),  # Manual ticket ID
        'client_name': data.get('client_name'),
        'subject': data.get('subject'),
        'description': data.get('description'),
        'status': 'open',
        'priority': data.get('priority', 'medium'),
        'user_id': str(current_user['_id']),
        'created_at': datetime.utcnow(),
        'updated_at': datetime.utcnow()
    }
    result = mongo.db.tickets.insert_one(ticket)
    ticket['_id'] = str(result.inserted_id)
    ticket['created_at'] = ticket['created_at'].isoformat()
    ticket['updated_at'] = ticket['updated_at'].isoformat()
    return jsonify(ticket), 201

@app.route('/api/tickets/<ticket_id>', methods=['GET'])
@token_required
def get_ticket(current_user, ticket_id):
    ticket = mongo.db.tickets.find_one({
        '_id': ObjectId(ticket_id),
        'user_id': str(current_user['_id'])
    })
    if not ticket:
        return jsonify({'message': 'Ticket not found'}), 404
    ticket['_id'] = str(ticket['_id'])
    ticket['created_at'] = ticket['created_at'].isoformat()
    if ticket.get('updated_at'):
        ticket['updated_at'] = ticket['updated_at'].isoformat()
    return jsonify(ticket), 200

@app.route('/api/tickets/<ticket_id>', methods=['PUT'])
@token_required
def update_ticket(current_user, ticket_id):
    try:
        data = request.get_json()
        print(f"Updating ticket {ticket_id} with data: {data}")
        
        ticket = mongo.db.tickets.find_one({
            '_id': ObjectId(ticket_id),
            'user_id': str(current_user['_id'])
        })
        if not ticket:
            print(f"Ticket {ticket_id} not found for user {current_user['_id']}")
            return jsonify({'message': 'Ticket not found'}), 404
        
        update_data = {'updated_at': datetime.utcnow()}
        for field in ['client_name', 'subject', 'description', 'status', 'priority']:
            if field in data:
                update_data[field] = data[field]
        
        print(f"Update data: {update_data}")
        result = mongo.db.tickets.update_one({'_id': ObjectId(ticket_id)}, {'$set': update_data})
        print(f"Update result: {result.modified_count} documents modified")
        
        return jsonify({'message': 'Ticket updated successfully'}), 200
    except Exception as e:
        print(f"Error updating ticket: {str(e)}")
        return jsonify({'message': f'Update failed: {str(e)}'}), 500

@app.route('/api/tickets/<ticket_id>', methods=['DELETE'])
@token_required
def delete_ticket(current_user, ticket_id):
    result = mongo.db.tickets.delete_one({
        '_id': ObjectId(ticket_id),
        'user_id': str(current_user['_id'])
    })
    if result.deleted_count == 0:
        return jsonify({'message': 'Ticket not found'}), 404
    mongo.db.ticket_comments.delete_many({'ticket_id': ticket_id})
    return jsonify({'message': 'Ticket deleted successfully'}), 200

# Ticket Comments
@app.route('/api/tickets/<ticket_id>/comments', methods=['GET'])
@token_required
def get_ticket_comments(current_user, ticket_id):
    comments = list(mongo.db.ticket_comments.find({'ticket_id': ticket_id}).sort('created_at', -1))
    for comment in comments:
        comment['_id'] = str(comment['_id'])
        comment['created_at'] = comment['created_at'].isoformat()
    return jsonify(comments), 200

@app.route('/api/tickets/<ticket_id>/comments', methods=['POST'])
@token_required
def add_ticket_comment(current_user, ticket_id):
    data = request.get_json()
    comment = {
        'ticket_id': ticket_id,
        'text': data.get('text'),
        'user_email': current_user['email'],
        'created_at': datetime.utcnow()
    }
    result = mongo.db.ticket_comments.insert_one(comment)
    comment['_id'] = str(result.inserted_id)
    comment['created_at'] = comment['created_at'].isoformat()
    return jsonify(comment), 201

# Export to PDF
@app.route('/api/export/pdf', methods=['GET'])
@token_required
def export_pdf(current_user):
    # Get date range parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Build query
    query = {'user_id': str(current_user['_id'])}
    if start_date or end_date:
        date_filter = {}
        if start_date:
            date_filter['$gte'] = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        if end_date:
            # Add one day to include the end date
            end_datetime = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            end_datetime = end_datetime.replace(hour=23, minute=59, second=59)
            date_filter['$lte'] = end_datetime
        query['created_at'] = date_filter
    
    todos = list(mongo.db.todos.find(query))
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title = "To Do App - Task List"
    if start_date or end_date:
        date_range = f" ({start_date or 'All'} to {end_date or 'All'})"
        title += date_range
    elements.append(Paragraph(title, styles['Title']))
    elements.append(Spacer(1, 20))
    
    # Table data
    data = [['Task', 'Category', 'Priority', 'Status', 'Created']]
    for todo in todos:
        status = '✓ Done' if todo.get('completed') else 'Pending'
        created = todo['created_at'].strftime('%Y-%m-%d') if isinstance(todo['created_at'], datetime) else str(todo['created_at'])[:10]
        data.append([
            todo.get('text', '')[:50],
            todo.get('category', 'personal'),
            todo.get('priority', 'medium'),
            status,
            created
        ])
    
    # Create table
    table = Table(data, colWidths=[200, 80, 70, 70, 80])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
    ]))
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name='todos.pdf')

# Export to Excel
@app.route('/api/export/excel', methods=['GET'])
@token_required
def export_excel(current_user):
    # Get date range parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Build query
    query = {'user_id': str(current_user['_id'])}
    if start_date or end_date:
        date_filter = {}
        if start_date:
            date_filter['$gte'] = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        if end_date:
            # Add one day to include the end date
            end_datetime = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            end_datetime = end_datetime.replace(hour=23, minute=59, second=59)
            date_filter['$lte'] = end_datetime
        query['created_at'] = date_filter
    
    todos = list(mongo.db.todos.find(query))
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Todos"
    
    # Add date range info if provided
    if start_date or end_date:
        ws.append([f"Export Date Range: {start_date or 'All'} to {end_date or 'All'}"])
        ws.append([])  # Empty row
    
    # Headers
    headers = ['Task', 'Category', 'Priority', 'Status', 'Created Date']
    ws.append(headers)
    
    # Style headers
    for col in range(1, 6):
        ws.cell(row=1, column=col).font = ws.cell(row=1, column=col).font.copy(bold=True)
    
    # Data
    for todo in todos:
        status = 'Completed' if todo.get('completed') else 'Pending'
        created = todo['created_at'].strftime('%Y-%m-%d') if isinstance(todo['created_at'], datetime) else str(todo['created_at'])[:10]
        ws.append([
            todo.get('text', ''),
            todo.get('category', 'personal'),
            todo.get('priority', 'medium'),
            status,
            created
        ])
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
                     as_attachment=True, download_name='todos.xlsx')

# Export Tickets to PDF
@app.route('/api/export/tickets/pdf', methods=['GET'])
@token_required
def export_tickets_pdf(current_user):
    # Get date range parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Build query
    query = {'user_id': str(current_user['_id'])}
    if start_date or end_date:
        date_filter = {}
        if start_date:
            date_filter['$gte'] = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        if end_date:
            # Add one day to include the end date
            end_datetime = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            end_datetime = end_datetime.replace(hour=23, minute=59, second=59)
            date_filter['$lte'] = end_datetime
        query['created_at'] = date_filter
    
    tickets = list(mongo.db.tickets.find(query))
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title = "To Do App - Tickets Report"
    if start_date or end_date:
        date_range = f" ({start_date or 'All'} to {end_date or 'All'})"
        title += date_range
    elements.append(Paragraph(title, styles['Title']))
    elements.append(Spacer(1, 20))
    
    # Table data
    data = [['Ticket ID', 'Client', 'Subject', 'Status', 'Priority', 'Created']]
    for ticket in tickets:
        created = ticket['created_at'].strftime('%Y-%m-%d') if isinstance(ticket['created_at'], datetime) else str(ticket['created_at'])[:10]
        data.append([
            ticket.get('ticket_id', '')[:15],
            ticket.get('client_name', '')[:20],
            ticket.get('subject', '')[:30],
            ticket.get('status', 'open'),
            ticket.get('priority', 'medium'),
            created
        ])
    
    # Create table
    table = Table(data, colWidths=[80, 100, 150, 70, 60, 80])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
    ]))
    elements.append(table)
    
    # Add comments for each ticket
    elements.append(Spacer(1, 30))
    elements.append(Paragraph("Ticket Details & Comments", styles['Heading2']))
    elements.append(Spacer(1, 15))
    
    for ticket in tickets:
        # Ticket header
        elements.append(Paragraph(f"<b>{ticket.get('ticket_id', 'N/A')}</b> - {ticket.get('subject', 'No Subject')}", styles['Heading3']))
        elements.append(Paragraph(f"Client: {ticket.get('client_name', 'N/A')} | Status: {ticket.get('status', 'open')} | Priority: {ticket.get('priority', 'medium')}", styles['Normal']))
        
        if ticket.get('description'):
            elements.append(Paragraph(f"Description: {ticket.get('description')}", styles['Normal']))
        
        # Get comments for this ticket
        comments = list(mongo.db.ticket_comments.find({'ticket_id': str(ticket['_id'])}).sort('created_at', 1))
        if comments:
            elements.append(Paragraph("Comments:", styles['Heading4']))
            for comment in comments:
                comment_date = comment['created_at'].strftime('%Y-%m-%d %H:%M') if isinstance(comment['created_at'], datetime) else str(comment['created_at'])[:16]
                elements.append(Paragraph(f"• {comment.get('text', '')} <i>({comment.get('user_email', 'Unknown')} - {comment_date})</i>", styles['Normal']))
        
        elements.append(Spacer(1, 15))
    
    doc.build(elements)
    buffer.seek(0)
    
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name='tickets.pdf')

# Export Tickets to Excel
@app.route('/api/export/tickets/excel', methods=['GET'])
@token_required
def export_tickets_excel(current_user):
    # Get date range parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Build query
    query = {'user_id': str(current_user['_id'])}
    if start_date or end_date:
        date_filter = {}
        if start_date:
            date_filter['$gte'] = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        if end_date:
            # Add one day to include the end date
            end_datetime = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            end_datetime = end_datetime.replace(hour=23, minute=59, second=59)
            date_filter['$lte'] = end_datetime
        query['created_at'] = date_filter
    
    tickets = list(mongo.db.tickets.find(query))
    
    wb = Workbook()
    
    # Tickets sheet
    ws = wb.active
    ws.title = "Tickets"
    
    # Add date range info if provided
    if start_date or end_date:
        ws.append([f"Export Date Range: {start_date or 'All'} to {end_date or 'All'}"])
        ws.append([])  # Empty row
    
    # Headers
    headers = ['Ticket ID', 'Client Name', 'Subject', 'Description', 'Status', 'Priority', 'Created Date']
    ws.append(headers)
    
    # Style headers
    for col in range(1, 8):
        ws.cell(row=1, column=col).font = ws.cell(row=1, column=col).font.copy(bold=True)
    
    # Data
    for ticket in tickets:
        created = ticket['created_at'].strftime('%Y-%m-%d') if isinstance(ticket['created_at'], datetime) else str(ticket['created_at'])[:10]
        ws.append([
            ticket.get('ticket_id', ''),
            ticket.get('client_name', ''),
            ticket.get('subject', ''),
            ticket.get('description', ''),
            ticket.get('status', 'open'),
            ticket.get('priority', 'medium'),
            created
        ])
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    
    # Comments sheet
    ws_comments = wb.create_sheet("Comments")
    comment_headers = ['Ticket ID', 'Comment', 'User', 'Date']
    ws_comments.append(comment_headers)
    
    # Style headers
    for col in range(1, 5):
        ws_comments.cell(row=1, column=col).font = ws_comments.cell(row=1, column=col).font.copy(bold=True)
    
    # Add comments data
    for ticket in tickets:
        comments = list(mongo.db.ticket_comments.find({'ticket_id': str(ticket['_id'])}).sort('created_at', 1))
        for comment in comments:
            comment_date = comment['created_at'].strftime('%Y-%m-%d %H:%M') if isinstance(comment['created_at'], datetime) else str(comment['created_at'])[:16]
            ws_comments.append([
                ticket.get('ticket_id', ''),
                comment.get('text', ''),
                comment.get('user_email', ''),
                comment_date
            ])
    
    # Adjust comment sheet column widths
    ws_comments.column_dimensions['A'].width = 15
    ws_comments.column_dimensions['B'].width = 50
    ws_comments.column_dimensions['C'].width = 25
    ws_comments.column_dimensions['D'].width = 20
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
                     as_attachment=True, download_name='tickets.xlsx')

@app.route('/', methods=['GET'])
def home():
    return jsonify({'message': 'Streamline API is running'}), 200

# For Vercel serverless
app = app

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
